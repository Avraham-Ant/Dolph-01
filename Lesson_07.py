"""
    本节课内容
    综合python 学习内容实现接口自动化测试  --- requests / openpyxl
"""

"""
    本节课基础知识内容
    主要是结合前两节课内容使用requests,openpyxl 实现对接口数据测试
    测试的步骤及内容具体如下：
    1.  用例读取测试数据 == 采用read_data
    2.  用数据发送接口请求 == post请求
    3.  得到执行结果与预期结果进行匹配 == if 
    4.  将得到的结论写入测试用例里面 == for 循环
    
    注意事项： 
    1.  eval 函数的使用  ==> 针对.dict.get方法输出的格式是字符串导致传输的数据格式不能通过的结果
    2.  replace 函数的使用 ==> 在进行eval转换的过程中python代码识别空为None，此时需要进行replace 函数替换(null===>None)否则报错
"""

import openpyxl  # 导入requests库、openpyxl库
import requests

session = requests.session()

# 构建与文件test_case.xlsx 测试用例模板的POST方法
def read_data():
    wd = openpyxl.load_workbook("test_case.xlsx")  # 载入相关文件"test_case.xlsx"
    ws = wd["recharge"]  # 确定导入数据的表
    list_read = []  # 创建一个空列表
    for i in range(2, (ws.max_row + 1)):  # ws.max_row 取得最大行,通过循环的方式 max_row

        dict_data = dict(  # 通过dict来构建字典形式以便在后续函数进行post请求时方便进行相关操作
            list_id=ws.cell(row=i, column=1).value,
            list_url=ws.cell(row=i, column=5).value,  # 获得表内url 地址 .cell(row=?,column=?)
            list_data=ws.cell(row=i, column=6).value,  # 获得表内数据data
            list_result=ws.cell(row=i, column=7).value,  # 获得表内实际结果
        )
        list_read.append(dict_data)
    return list_read


results = read_data()  # 获得返回值 取得相对应的结果
# print(results)

# print("通过read_data函数方法获得的数据为：", results)
# 构建相关post方法进行数据的传输 <=== 将拿到的result作为参数传递进去
def data_post():
    bool_list = []  # 先定义一个空列表来存储对比数据

    for result in results:  # 将传输的数据利用循环分别存入id,url,data
        post_id = result.get("list_id")  # 获得传输数据里的list_id
        post_url = result.get("list_url")  # 获得传输数据里的list_url
        post_data = result.get("list_data")  # 获得传输数据里的list_data
        post_result = result["list_result"]  # 获得传输数据的实际结果
        print("post_result", post_result, type(post_result))
        post_result = post_result.replace('null', 'None')
        post_result = eval(post_result)  # 报错原因：次数null 必须 替换成 none 将post_result(理想结果)进行eval函数的转换
        # print("打印获取的相关信息：", post_id, post_data, post_url, post_result)

        #   eval 函数的使用 str ---> dict
        post_data = eval(post_data)

        res = session.post(url=post_url,
                            data=post_data)  # 采用post 内部方法进行相对于requests.post(url, data=None, json=None, **kwargs):

        real_response = res.json()  # 返回值需要采用json 格式故这里解码需要响应
        print(real_response)  # 传输过程中会有报错 == 传输的数据格式有问题 == 因为get 方法取出来的格式是string 而传输数据是字典的格式故应该进行eval 的转换 以及 相关null
        # --> none
        #   获得数据后要将实际结果和理想结果进行对比 ==== 由于post_result 也是get方法获得的格式：str;而实际结果是字典的形式此处也需要进行eval
        #   将获得的真实结果requests.post 与其进行比较 == 采用if函数
        if post_result == real_response:
            bool_list.append('PASS')
        else:
            bool_list.append('FAiled')
        print("bool_list:", bool_list)

    return bool_list


write_data = data_post()  # 讲需要写入的数据用write_data 进行存储方便写入


#   下面进行写入数据的相关操作data_write方法的写入 === openpyxl方法的使用
def data_write():
    write_wd = openpyxl.load_workbook("test_case.xlsx")  # 载入相关文件"test_case.xlsx"
    write_ws = write_wd["recharge"]  # 确定导入数据的表
    # 利用循环讲获得的write_data 依次写入文件实际结果部位
    for i in range(2, write_ws.max_row +1):
        write_ws.cell(row=i, column=8).value = write_data[i-2]
    write_wd.save("test_case.xlsx")

data_write()

#   在搭建完基本程序后需要实现函数内以及函数间一些变量的声明，不固定相对于的参数 例如 文件名 / 列表名

